from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app, Response
from flask_login import current_user, login_required
from models import db, Collection, Product, ProductColorVariant, Design
from utils.mockups import get_carousel_colors_for_product
from sqlalchemy.orm import joinedload
import json
import io

collection_bp = Blueprint('collection', __name__, url_prefix='/c')

@collection_bp.route('/<slug>')
def view(slug):
    """View collection landing page - design board of available items"""
    collection = Collection.query.options(
        joinedload(Collection.products)
    ).filter_by(slug=slug, is_active=True).first_or_404()
    
    # Check password if protected
    if collection.is_password_protected:
        if not session.get(f'collection_{collection.id}_access'):
            return redirect(url_for('collection.password', slug=slug))
    
    # Set collection in session for checkout
    session['collection_id'] = collection.id
    
    # Get allowed colors when organizer restricted options
    allowed_colors = None
    if collection.restrict_options and collection.allowed_colors:
        allowed_colors = set(json.loads(collection.allowed_colors))
    
    # Get products in this collection with carousel colors (DB + mockup folder)
    all_products = collection.products
    products = []
    unavailable_products = []  # Products that don't have any of the chosen colors
    for product in all_products:
        variants = get_carousel_colors_for_product(product, current_app, allowed_colors=allowed_colors)
        product.carousel_colors = variants
        product.available_sizes_list = json.loads(product.available_sizes) if product.available_sizes else []
        # When colors are restricted: only show products that have at least one matching color
        if allowed_colors:
            if variants:
                products.append(product)
            else:
                unavailable_products.append(product.name)
        else:
            products.append(product)
    
    return render_template('collection/view.html', 
                         collection=collection,
                         products=products,
                         unavailable_products=unavailable_products)


@collection_bp.route('/<slug>/password', methods=['GET', 'POST'])
def password(slug):
    """Password protection for collection"""
    collection = Collection.query.filter_by(slug=slug, is_active=True).first_or_404()
    
    if not collection.is_password_protected:
        return redirect(url_for('collection.view', slug=slug))
    
    if request.method == 'POST':
        password = request.form.get('password')
        
        if collection.check_password(password):
            session[f'collection_{collection.id}_access'] = True
            return redirect(url_for('collection.view', slug=slug))
        else:
            flash('Incorrect password', 'error')
    
    return render_template('collection/password.html', collection=collection)


@collection_bp.route('/<slug>/share')
def share(slug):
    """Collection share page (shows share link and QR code)"""
    collection = Collection.query.filter_by(slug=slug, is_active=True).first_or_404()

    # Check password if protected
    if collection.is_password_protected:
        if not session.get(f'collection_{collection.id}_access'):
            return redirect(url_for('collection.password', slug=slug))

    # Resolve designs for this collection so the share page can show them
    designs = []
    if collection.allowed_design_ids:
        try:
            ids = json.loads(collection.allowed_design_ids)
            if ids:
                designs = Design.query.filter(Design.id.in_(ids)).all()
        except Exception:
            designs = []

    # Determine if the current user can manage (delete) designs
    can_manage = (
        current_user.is_authenticated and (
            getattr(current_user, 'is_admin', False) or
            current_user.id == collection.created_by_user_id
        )
    )

    return render_template('collection/share.html', collection=collection,
                           designs=designs, can_manage=can_manage)


@collection_bp.route('/<slug>/design/<int:design_id>/delete', methods=['POST'])
@login_required
def delete_design(slug, design_id):
    """Remove a design from a group order's allowed list.

    Only the collection creator or an admin may do this.
    """
    collection = Collection.query.filter_by(slug=slug).first_or_404()

    # Permission check
    is_admin = getattr(current_user, 'is_admin', False)
    is_creator = current_user.id == collection.created_by_user_id
    if not (is_admin or is_creator):
        flash('You do not have permission to delete designs from this group order.', 'error')
        return redirect(url_for('collection.share', slug=slug))

    # Remove the design from allowed_design_ids
    try:
        ids = json.loads(collection.allowed_design_ids) if collection.allowed_design_ids else []
        ids = [i for i in ids if i != design_id]
        collection.allowed_design_ids = json.dumps(ids) if ids else None
        db.session.commit()
        flash('Design removed from this group order.', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error removing design %s from collection %s: %s', design_id, slug, e)
        flash('Could not remove the design. Please try again.', 'error')

    return redirect(url_for('collection.share', slug=slug))


# ── Excel Export ─────────────────────────────────────────────────────────────

@collection_bp.route('/<slug>/export.xlsx')
@login_required
def export_xlsx(slug):
    """Download an Excel workbook for this group order.

    Accessible to:
      - Site admins (is_admin flag)
      - The user who created the collection (created_by_user_id)
    """
    collection = Collection.query.filter_by(slug=slug).first_or_404()

    is_admin   = getattr(current_user, 'is_admin', False)
    is_creator = (current_user.id == collection.created_by_user_id)
    if not (is_admin or is_creator):
        flash('You do not have permission to export this group order.', 'error')
        return redirect(url_for('collection.share', slug=slug))

    try:
        xlsx_bytes = _build_group_order_xlsx(collection)
    except Exception as e:
        current_app.logger.exception('Excel export failed for collection %s: %s', slug, e)
        flash('Could not generate the Excel file. Please try again.', 'error')
        return redirect(url_for('collection.share', slug=slug))

    safe_name = slug.replace('/', '_').replace(' ', '_')
    filename  = f"group_order_{safe_name}.xlsx"
    return Response(
        xlsx_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _build_group_order_xlsx(collection):
    """Build and return raw .xlsx bytes for the given collection."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from models import Order, OrderItem

    # ── colour palette ──────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill('solid', fgColor='B87C6F')   # terra-cotta
    HEADER_FONT  = Font(bold=True, color='FFFFFF', size=11)
    TITLE_FONT   = Font(bold=True, size=14)
    THIN         = Side(style='thin', color='CCCCCC')
    THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    MONEY_FMT    = '$#,##0.00'
    CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT         = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.border    = THIN_BORDER
            cell.alignment = CENTER

    def style_data_row(ws, row, cols, alt=False):
        fill = PatternFill('solid', fgColor='FAF7F4') if alt else PatternFill('solid', fgColor='FFFFFF')
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = LEFT

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Summary
    # ════════════════════════════════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = 'Summary'

    orders = collection.orders.order_by(Order.created_at.asc()).all()

    from datetime import datetime
    ws_sum['A1'] = f'Group Order: {collection.name}'
    ws_sum['A1'].font = TITLE_FONT
    ws_sum.merge_cells('A1:F1')
    ws_sum['A1'].alignment = CENTER

    ws_sum['A2'] = f'Exported: {datetime.utcnow().strftime("%B %d, %Y %H:%M UTC")}'
    ws_sum['A2'].font = Font(italic=True, color='888888', size=9)
    ws_sum.merge_cells('A2:F2')

    ws_sum['A3'] = f'Deadline: {collection.order_deadline.strftime("%B %d, %Y") if collection.order_deadline else "—"}'
    ws_sum.merge_cells('A3:F3')

    paid_orders   = [o for o in orders if o.payment_status == 'paid']
    total_revenue = sum(o.total for o in paid_orders)
    total_items   = sum(item.quantity for o in orders for item in o.items)

    ws_sum['A5'] = 'Total Orders';  ws_sum['B5'] = len(orders)
    ws_sum['A6'] = 'Paid Orders';   ws_sum['B6'] = len(paid_orders)
    ws_sum['A7'] = 'Total Revenue'; ws_sum['B7'] = total_revenue; ws_sum['B7'].number_format = MONEY_FMT
    ws_sum['A8'] = 'Total Items';   ws_sum['B8'] = total_items

    for row in range(5, 9):
        ws_sum.cell(row=row, column=1).font = Font(bold=True)

    ws_sum.row_dimensions[1].height = 28
    ws_sum.column_dimensions['A'].width = 20
    ws_sum.column_dimensions['B'].width = 18

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Orders
    # ════════════════════════════════════════════════════════════════════════
    ws_ord = wb.create_sheet('Orders')
    ord_headers = [
        'Order #', 'Date', 'Name', 'Email', 'Phone',
        'Fulfillment', 'Payment', 'Status',
        'Items', 'Subtotal', 'Shipping', 'Tax', 'Total', 'Notes',
    ]
    ws_ord.append(ord_headers)
    style_header_row(ws_ord, 1, len(ord_headers))
    ws_ord.row_dimensions[1].height = 22

    for idx, order in enumerate(orders, start=2):
        items_count = sum(i.quantity for i in order.items)
        ws_ord.append([
            order.order_number,
            order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else '',
            order.full_name,
            order.email or '',
            order.phone or '',
            order.fulfillment_method or '',
            order.payment_status or '',
            order.status or '',
            items_count,
            order.subtotal,
            order.shipping_cost or 0,
            order.tax or 0,
            order.total,
            order.customer_notes or '',
        ])
        style_data_row(ws_ord, idx, len(ord_headers), alt=(idx % 2 == 0))
        for col in (10, 11, 12, 13):
            ws_ord.cell(row=idx, column=col).number_format = MONEY_FMT

    set_col_widths(ws_ord, [18, 16, 22, 28, 14, 12, 10, 14, 7, 10, 10, 8, 10, 30])

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 3 — Line Items
    # ════════════════════════════════════════════════════════════════════════
    ws_items = wb.create_sheet('Line Items')
    item_headers = [
        'Order #', 'Customer', 'Product', 'Style #',
        'Color', 'Size', 'Qty', 'Placement',
        'Design', 'Back (Name/Number)', 'Unit Price', 'Line Total',
    ]
    ws_items.append(item_headers)
    style_header_row(ws_items, 1, len(item_headers))
    ws_items.row_dimensions[1].height = 22

    row_idx = 2
    for order in orders:
        for item in order.items:
            back_meta = {}
            if item.back_design_meta:
                try:
                    back_meta = json.loads(item.back_design_meta)
                except Exception:
                    pass
            back_str = ''
            if back_meta:
                parts = []
                if back_meta.get('name'):
                    parts.append(f"Name: {back_meta['name']}")
                if back_meta.get('number'):
                    parts.append(f"#{back_meta['number']}")
                back_str = '  |  '.join(parts)

            ws_items.append([
                order.order_number,
                order.full_name,
                item.product_name or '',
                item.style_number or '',
                item.color or '',
                item.size or '',
                item.quantity,
                item.placement or '',
                item.design_file_name or '',
                back_str,
                item.unit_price,
                item.subtotal,
            ])
            style_data_row(ws_items, row_idx, len(item_headers), alt=(row_idx % 2 == 0))
            ws_items.cell(row=row_idx, column=11).number_format = MONEY_FMT
            ws_items.cell(row=row_idx, column=12).number_format = MONEY_FMT
            row_idx += 1

    set_col_widths(ws_items, [18, 22, 28, 10, 18, 7, 5, 14, 26, 22, 10, 10])

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 4 — Size Breakdown (production tally)
    # ════════════════════════════════════════════════════════════════════════
    ws_sizes = wb.create_sheet('Size Breakdown')

    SIZE_ORDER = ['NB', '3M', '6M', '12M', '18M', '24M',
                  '2T', '3T', '4T', '5T',
                  'YXS', 'YS', 'YM', 'YL', 'YXL',
                  'XS', 'S', 'M', 'L', 'XL', '2XL', '3XL', '4XL', '5XL']

    tally    = {}
    all_sizes = set()
    for order in orders:
        for item in order.items:
            key  = (item.product_name or '', item.color or '')
            size = item.size or '?'
            all_sizes.add(size)
            tally.setdefault(key, {})
            tally[key][size] = tally[key].get(size, 0) + item.quantity

    known_set   = set(SIZE_ORDER)
    extra_sizes = sorted(s for s in all_sizes if s not in known_set)
    sorted_sizes = [s for s in SIZE_ORDER if s in all_sizes] + extra_sizes

    size_hdr = ['Product', 'Color'] + sorted_sizes + ['TOTAL']
    ws_sizes.append(size_hdr)
    style_header_row(ws_sizes, 1, len(size_hdr))
    ws_sizes.row_dimensions[1].height = 22

    row_idx = 2
    for (product_name, color), size_map in sorted(tally.items()):
        row_data  = [product_name, color]
        row_total = 0
        for size in sorted_sizes:
            qty = size_map.get(size, 0)
            row_total += qty
            row_data.append(qty if qty else '')
        row_data.append(row_total)
        ws_sizes.append(row_data)
        style_data_row(ws_sizes, row_idx, len(size_hdr), alt=(row_idx % 2 == 0))
        ws_sizes.cell(row=row_idx, column=len(size_hdr)).font = Font(bold=True)
        row_idx += 1

    # Grand total row
    gt_row    = ['', 'TOTAL']
    grand_total = 0
    for size in sorted_sizes:
        col_total = sum(tally[k].get(size, 0) for k in tally)
        gt_row.append(col_total if col_total else '')
        grand_total += col_total
    gt_row.append(grand_total)
    ws_sizes.append(gt_row)
    for col in range(1, len(size_hdr) + 1):
        cell = ws_sizes.cell(row=row_idx, column=col)
        cell.fill   = PatternFill('solid', fgColor='3A3D48')
        cell.font   = Font(bold=True, color='FFFFFF')
        cell.border = THIN_BORDER

    set_col_widths(ws_sizes, [28, 18] + [6] * len(sorted_sizes) + [8])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
